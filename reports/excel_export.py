"""
Generates an Excel compliance report from a ComplianceReport dict.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

GREEN_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
RED_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
HEADER_FILL = PatternFill(start_color="0B2545", end_color="0B2545", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def generate_excel_report(report: dict, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compliance Report"

    ws["A1"] = "GeM Bid Compliance Report"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Bid ID: {report['bid_id']}"
    ws["A3"] = f"Overall Status: {report['overall_status']}"
    ws["A3"].font = Font(bold=True)

    headers = ["Document Type", "File Name", "Field", "Status", "Reason"]
    header_row = 5
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    row = header_row + 1
    for doc in report["documents"]:
        for result in doc["results"]:
            ws.cell(row=row, column=1, value=doc["doc_type"])
            ws.cell(row=row, column=2, value=doc.get("file_name") or "Not submitted")
            ws.cell(row=row, column=3, value=result["field"])
            status_cell = ws.cell(row=row, column=4, value="Pass" if result["passed"] else "Fail")
            ws.cell(row=row, column=5, value=result["reason"] or "-")
            status_cell.fill = GREEN_FILL if result["passed"] else RED_FILL
            row += 1

    for col_letter, width in zip("ABCDE", [16, 24, 16, 10, 40]):
        ws.column_dimensions[col_letter].width = width

    wb.save(output_path)


if __name__ == "__main__":
    import sys, os
    sys.path.append(os.path.join(os.path.dirname(__file__), ".."))
    from extraction.pdf_reader import extract_text_from_pdf
    from extraction.field_extractor import process_document
    from rules.rule_engine import load_rules, build_compliance_report

    folder = "data/sample_docs/valid"
    rules = load_rules()
    extracted_docs = []
    for filename in os.listdir(folder):
        path = os.path.join(folder, filename)
        text = extract_text_from_pdf(path)
        extracted_docs.append(process_document(path, text))

    report = build_compliance_report("BID-001", extracted_docs, rules)
    generate_excel_report(report, "reports/sample_compliance_report.xlsx")
    print("Excel report generated: reports/sample_compliance_report.xlsx")